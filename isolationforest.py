import pandas as pd
import numpy as np
from sklearn.ensemble import IsolationForest
import openpyxl
from openpyxl import load_workbook
rng = np.random.RandomState(42)
df=pd.read_excel('D:/AnacondaSpyder/300403iso.xlsx')
data=df.values

# fit the model
clf = IsolationForest(max_samples=256, random_state=rng)
clf.fit(data)
y_pred_outliers = clf.predict(data)
#print(y_pred_outliers)

def write_excel_xlsx(path, sheet_name, values):
    index = len(values)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
            sheet.cell(row=i+1, column=1, value=values[i])
    workbook.save(path)
    print("xlsx格式表格写入数据成功！")
 
 
def read_excel_xlsx(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    # sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
    sheet = workbook[sheet_name]
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()
 
book_name_xlsx = '异常值表格300403.xlsx'
 
sheet_name_xlsx = '异常值表'
 
write_excel_xlsx(book_name_xlsx, sheet_name_xlsx, y_pred_outliers)
#read_excel_xlsx(book_name_xlsx, sheet_name_xlsx)


